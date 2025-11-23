import os
import re
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import random
from collections import defaultdict, OrderedDict


GENOTYPE_ORDER = ["IL-17 KO", "C57Bl/6J"]


# ------------------ CORE LOGIC ------------------ #

def assign_balanced_exit_arms(animal_data):
    """
    Learning-days exit-arm assignment:
      1) Within each (Genotype, Sex, Cage) group we 'cycle' arms for fairness.
      2) Globally, totals for arms 1/2/3 are as balanced as possible.
    """
    from collections import Counter

    # Group animals
    groupings = defaultdict(list)
    for animal in animal_data:
        key = (animal['Genotype'], animal['Sex'], animal['Cage'])
        groupings[key].append(animal)

    # Shuffle for fairness (seed is controlled by GUI)
    groups = []
    for key, animals in groupings.items():
        animals = animals[:]
        random.shuffle(animals)
        groups.append((key, animals))
    random.shuffle(groups)

    # Global targets
    N = len(animal_data)
    base = N // 3
    remainder = N % 3
    target = {1: base + (1 if remainder >= 1 else 0),
              2: base + (1 if remainder >= 2 else 0),
              3: base}
    remaining = target.copy()

    def seq_from_offset(m, offset):
        return [((j + offset) % 3) + 1 for j in range(m)]

    def score_sequence(seq, rem):
        c = Counter(seq)
        over = sum(max(0, c[a] - rem[a]) for a in (1, 2, 3))
        prefer = -sum(min(c[a], rem[a]) for a in (1, 2, 3))
        return (over, prefer)

    exit_arm_map = {}
    for _, animals in groups:
        m = len(animals)
        candidates = [(o, seq_from_offset(m, o)) for o in (0, 1, 2)]
        best_offset, best_seq = min(candidates, key=lambda it: score_sequence(it[1], remaining))

        adjusted = []
        for arm in best_seq:
            if remaining[arm] > 0:
                adjusted.append(arm); remaining[arm] -= 1
            else:
                alt = max((1, 2, 3), key=lambda a: (remaining[a], random.random()))
                adjusted.append(alt); remaining[alt] -= 1

        for animal, arm in zip(animals, adjusted):
            exit_arm_map[animal['AnimalID']] = arm

    return exit_arm_map


def _norm_hyphens(s: str) -> str:
    # normalize various hyphens to ASCII hyphen-minus
    return (s.replace("\u2011", "-")  # non-breaking hyphen
             .replace("\u2013", "-")  # en dash
             .replace("\u2014", "-")  # em dash
             .replace("\u2212", "-")  # minus
             .strip())


def group_animals_by_cage_in_order(animal_data):
    """
    Preserve first-seen cage order; keep animal order within each cage.
    Returns OrderedDict[cage] -> list[animal_dict]
    """
    cages = OrderedDict()
    for a in animal_data:
        cages.setdefault(a['Cage'], []).append(a)
    return cages


def plan_cage_dp(cage_animals, prev_arm, learning_exit_map):
    r"""
    Dynamic program INSIDE one cage (animals in fixed order).
    For each animal i, allowed colors = {1,2,3} \ {learning_exit[i]}  (set difference; always size 2).
    Cost = number of switches between consecutive assigned arms,
           plus a boundary switch if first != prev_arm.
    Returns dict: {'cost', 'end', 'first', 'colors'}
    """
    m = len(cage_animals)
    if m == 0:
        return {'cost': 0, 'end': prev_arm, 'first': prev_arm, 'colors': []}

    # allowed per position
    allowed = []
    for a in cage_animals:
        forbid = learning_exit_map[a['AnimalID']]
        allowed.append(tuple(sorted({1, 2, 3} - {forbid})))  # two arms

    dp = [dict() for _ in range(m)]      # dp[i][c] = (cost, prev_color)
    # first position
    for c in allowed[0]:
        cost0 = 0 if (prev_arm is None or c == prev_arm) else 1
        dp[0][c] = (cost0, None)

    # propagate
    for i in range(1, m):
        for c in allowed[i]:
            best_cost = None
            best_prev = None
            for c_prev, (cost_prev, _) in dp[i-1].items():
                cost_here = cost_prev + (0 if c == c_prev else 1)
                if best_cost is None or cost_here < best_cost:
                    best_cost = cost_here
                    best_prev = c_prev
            dp[i][c] = (best_cost, best_prev)

    # choose end with min cost
    end_color = min(dp[-1], key=lambda c: dp[-1][c][0])
    total_cost = dp[-1][end_color][0]

    # reconstruct sequence
    colors = [None] * m
    colors[-1] = end_color
    for i in range(m-1, 0, -1):
        colors[i-1] = dp[i][colors[i]][1]
    first_color = colors[0]

    return {'cost': total_cost, 'end': end_color, 'first': first_color, 'colors': colors}



def build_nonlearning_plan_cage_packs(animal_data, learning_exit_map):
    """
    NON-LEARNING days:
      - Reorder *cages* (packs), never split a cage.
      - Keep animals within each cage in original order.
      - Assign exits per animal (≠ learning exit) to minimize total switches,
        including wrap-around.
      - Greedy over all start-arms (1/2/3) × first-cage choices, using
        per-cage DP that accounts for the incoming arm.
    Returns:
      ordered_animals: list of animal dicts in final cage-pack order
      per_animal_exit: dict AnimalID -> assigned exit for the day
    """
    cages = group_animals_by_cage_in_order(animal_data)
    cage_names = list(cages.keys())

    # Precompute per-cage plans for every incoming arm
    plans = {c: {s: plan_cage_dp(cages[c], s, learning_exit_map) for s in (1, 2, 3)}
             for c in cage_names}

    best_total = None
    best_solution = None  # (start_arm, cage_sequence, chosen_plans_per_cage)

    for start_arm in (1, 2, 3):
        for first_cage in cage_names:
            remaining = set(cage_names)
            seq = []
            chosen = {}
            prev_end = start_arm
            total_cost = 0

            current = first_cage
            while remaining:
                if current is None:
                    # pick next cage that is cheapest given prev_end
                    current = min(remaining, key=lambda ck: plans[ck][prev_end]['cost'])
                plan = plans[current][prev_end]
                total_cost += plan['cost']
                seq.append(current)
                chosen[current] = plan
                remaining.remove(current)
                prev_end = plan['end']
                current = None

            # wrap-around penalty (cycle)
            total_cost += 0 if prev_end == start_arm else 1

            if best_total is None or total_cost < best_total:
                best_total = total_cost
                best_solution = (start_arm, seq, chosen)

    # Build final output
    _, cage_sequence, chosen_plans = best_solution
    per_animal_exit = {}
    ordered_animals = []

    for c in cage_sequence:
        animals = cages[c]
        colors = chosen_plans[c]['colors']
        for a, arm in zip(animals, colors):
            per_animal_exit[a['AnimalID']] = arm
            ordered_animals.append(a)

    return ordered_animals, per_animal_exit


def generate_pseudorandom_sequence(n_trials, armA, armB, avoid_first=None):
    """
    Generate a pseudo-random start-arm sequence (values armA/armB) with near-balance
    and no triple repeats. If avoid_first is provided, first element != avoid_first.
    """
    countA, countB = n_trials // 2, n_trials // 2
    if n_trials % 2 == 1:
        if random.random() < 0.5: countA += 1
        else: countB += 1

    pool = [armA] * countA + [armB] * countB
    random.shuffle(pool)

    def valid(seq): return len(seq) < 3 or not (seq[-1] == seq[-2] == seq[-3])

    def backtrack(cur, items):
        if not items: return cur
        for i, cand in enumerate(items):
            if avoid_first is not None and len(cur) == 0 and cand == avoid_first:
                continue
            new = cur + [cand]
            if valid(new):
                out = backtrack(new, items[:i] + items[i+1:])
                if out is not None: return out
        return None

    seq = backtrack([], pool)
    return seq if seq else pool


def generate_day_tables(animal_data, learning_days, reversal_days, n_trials, exit_arm_map):
    """
    Build per-day tables.

    Learning days:
      - Use 'exit_arm_map' (balanced) and keep animal order as entered.

    Non-learning days (reversal days):
      - Reorder *cages only* (packs intact), keep animal order within each cage.
      - Assign exits per animal to minimize total exit-arm switches (cyclic).
      - Each animal’s non-learning exit ≠ its learning-day exit.
    """
    total_days = learning_days + reversal_days
    if total_days <= 0:
        return [], []

    base_order_animals = list(animal_data)

    day_tables_data, day_tables_text = [], []

    for d in range(total_days):
        in_learning = d < learning_days
        header = ["AnimalID", "Tag", "Sex", "Genotype", "Cage", "ExitArm"] + [f"T{i+1}" for i in range(n_trials)]
        rows = []

        if in_learning:
            # keep original order; exits are the balanced learning exits
            animals_today = base_order_animals
            per_day_exit = {a['AnimalID']: exit_arm_map[a['AnimalID']] for a in animals_today}
        else:
            # reorder cages + per-animal assignment to minimize switches
            animals_today, per_day_exit = build_nonlearning_plan_cage_packs(base_order_animals, exit_arm_map)

        for animal in animals_today:
            aid = animal['AnimalID']
            exit_arm_today = per_day_exit[aid]

            # start-arm choices exclude today's exit
            other = [1, 2, 3]
            other.remove(exit_arm_today)
            armA, armB = other

            # Learning days: avoid_first = today's ExitArm
            # Non-learning: avoid_first = learning-day exit (differs from prior phase)
            avoid_first = exit_arm_today if in_learning else exit_arm_map[aid]

            seq = generate_pseudorandom_sequence(n_trials, armA, armB, avoid_first=avoid_first)
            row = [aid, animal['Tag'], animal['Sex'], animal['Genotype'], animal['Cage'], exit_arm_today] + seq
            rows.append(row)

        text_lines = [f"Day {d+1} ({'Learning' if in_learning else 'Reversal'}):"]
        text_lines.append("\t".join(header))
        for r in rows:
            text_lines.append("\t".join(str(x) for x in r))
        table_str = "\n".join(text_lines)

        day_tables_data.append([header] + rows)
        day_tables_text.append(table_str)

    return day_tables_data, day_tables_text


# ------------------ TKINTER GUI ------------------ #

class YMazeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Shallow Water Y-Maze Scheduler")
        self.day_tables_data = None
        self.day_tables_text = None
        self._build_ui()

    # ---------- UI ---------- #
    def _build_ui(self):
        main = ttk.Frame(self, padding=10)
        main.grid(sticky="nsew")
        self.columnconfigure(0, weight=1)
        self.rowconfigure(0, weight=1)

        lbl_info = ttk.Label(
            main,
            text=("Paste Animal Data (tab- or 2+ space-separated columns).\n"
                  "Cols: AnimalID | Tag | Sex | Genotype | ... | Cage (last)")
        )
        lbl_info.grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 6))

        self.txt_animals = tk.Text(main, width=110, height=12, font=("Consolas", 10))
        self.txt_animals.grid(row=1, column=0, columnspan=6, sticky="nsew")

        # Params
        ttk.Label(main, text="Learning days:").grid(row=2, column=0, sticky="e")
        self.var_learning = tk.StringVar(value="3")
        ttk.Entry(main, textvariable=self.var_learning, width=6).grid(row=2, column=1, sticky="w")

        ttk.Label(main, text="Reversal days:").grid(row=2, column=2, sticky="e")
        self.var_reversal = tk.StringVar(value="2")
        ttk.Entry(main, textvariable=self.var_reversal, width=6).grid(row=2, column=3, sticky="w")

        ttk.Label(main, text="Trials/day:").grid(row=2, column=4, sticky="e")
        self.var_trials = tk.StringVar(value="10")
        ttk.Entry(main, textvariable=self.var_trials, width=6).grid(row=2, column=5, sticky="w")

        # RNG seed
        self.var_seed_on = tk.BooleanVar(value=False)
        self.var_seed_val = tk.StringVar(value="42")
        ttk.Checkbutton(main, text="Seed RNG", variable=self.var_seed_on).grid(row=3, column=0, sticky="e")
        ttk.Entry(main, textvariable=self.var_seed_val, width=8).grid(row=3, column=1, sticky="w")

        # Buttons
        ttk.Button(main, text="Generate Tables", command=self.on_generate).grid(row=3, column=2, padx=4, pady=6, sticky="w")
        ttk.Button(main, text="Copy Output", command=self.on_copy).grid(row=3, column=3, padx=4, pady=6, sticky="w")

        # Export buttons
        ttk.Button(main, text="Export → CSVs (one per day)", command=self.on_export_csvs).grid(row=3, column=4, padx=4, pady=6, sticky="w")
        ttk.Button(main, text="Export → Combined CSV", command=self.on_export_one_csv).grid(row=3, column=5, padx=4, pady=6, sticky="w")
        ttk.Button(main, text="Export → Excel (sheets)", command=self.on_export_xlsx).grid(row=4, column=4, padx=4, pady=0, sticky="w")

        # Output
        ttk.Label(main, text="Output:").grid(row=4, column=0, sticky="w", pady=(8, 2))
        self.txt_output = tk.Text(main, width=110, height=22, font=("Consolas", 10))
        self.txt_output.grid(row=5, column=0, columnspan=6, sticky="nsew")

        # Status bar
        self.var_status = tk.StringVar(value="Ready.")
        ttk.Label(main, textvariable=self.var_status).grid(row=6, column=0, columnspan=6, sticky="w", pady=(6, 0))

        # Resizing
        main.columnconfigure(0, weight=0)
        main.columnconfigure(1, weight=0)
        main.columnconfigure(2, weight=0)
        main.columnconfigure(3, weight=1)
        main.columnconfigure(4, weight=0)
        main.columnconfigure(5, weight=0)
        main.rowconfigure(1, weight=0)
        main.rowconfigure(5, weight=1)

    # ---------- Parsing ---------- #
    @staticmethod
    def _smart_split(line: str):
        if "\t" in line:
            parts = [p.strip() for p in line.split("\t") if p.strip() != ""]
        else:
            parts = [p.strip() for p in re.split(r"\s{2,}", line) if p.strip() != ""]
        return parts

    def parse_animal_data(self, text):
        """
        Robust parser: AnimalID | [Tag...] | Sex | [Genotype...] | Cage(last)
        Anchored by the Sex token.
        """
        lines = [line for line in text.splitlines() if line.strip()]
        animal_data = []
        for raw in lines:
            line = raw.strip()
            if not line:
                continue
            if "AnimalID" in line and "Sex" in line:
                continue

            tokens = line.split()
            if len(tokens) < 4:
                continue

            sex_idx = None
            for i, t in enumerate(tokens):
                if t in ("Male", "Female"):
                    sex_idx = i
                    break
            if sex_idx is None or sex_idx < 1 or len(tokens) < sex_idx + 2:
                continue

            animal_id = tokens[0]
            tag = " ".join(tokens[1:sex_idx]).strip()
            sex = tokens[sex_idx]
            cage = tokens[-1]
            genotype = " ".join(tokens[sex_idx + 1:-1]).strip()

            d = {
                'AnimalID': _norm_hyphens(animal_id),
                'Tag': _norm_hyphens(tag),
                'Sex': _norm_hyphens(sex),
                'Genotype': _norm_hyphens(genotype),
                'Cage': _norm_hyphens(cage),
            }
            if not d['Genotype'] or not d['Cage']:
                continue
            animal_data.append(d)
        return animal_data

    # ---------- Actions ---------- #
    def on_generate(self):
        try:
            learning_days = int(self.var_learning.get())
            reversal_days = int(self.var_reversal.get())
            n_trials = int(self.var_trials.get())
        except ValueError:
            messagebox.showerror("Invalid input", "Learning days, reversal days, and trials must be integers.")
            return

        raw_text = self.txt_animals.get("1.0", tk.END)
        animal_data = self.parse_animal_data(raw_text)
        if not animal_data:
            self._set_status("No valid animal rows parsed. Check spacing/tabs.")
            self._write_output("No valid animal rows parsed. Check spacing/tabs.\n")
            return

        if learning_days < 0 or reversal_days < 0 or n_trials <= 0:
            messagebox.showerror("Invalid input", "Days must be >= 0 and trials/day must be > 0.")
            return

        if self.var_seed_on.get():
            try:
                random.seed(int(self.var_seed_val.get()))
            except ValueError:
                messagebox.showerror("Invalid seed", "Seed must be an integer.")
                return
        else:
            random.seed()

        # Learning-day baseline map
        learning_exit_map = assign_balanced_exit_arms(animal_data)

        self.day_tables_data, self.day_tables_text = generate_day_tables(
            animal_data, learning_days, reversal_days, n_trials, learning_exit_map
        )

        self.txt_output.delete("1.0", tk.END)
        for t in self.day_tables_text:
            self.txt_output.insert(tk.END, t + "\n\n")

        self._set_status(f"Generated {learning_days + reversal_days} day(s) × {n_trials} trials.")

    def on_copy(self):
        if not self.day_tables_text:
            self._set_status("Nothing to copy. Generate tables first.")
            return
        self.clipboard_clear()
        self.clipboard_append("\n\n".join(self.day_tables_text))
        self._set_status("Copied output to clipboard.")

    def on_export_csvs(self):
        if not self.day_tables_data:
            self._write_output("Nothing to export. Generate tables first.\n")
            self._set_status("Export aborted: nothing to export.")
            return
        folder = filedialog.askdirectory(title="Choose folder to save CSVs")
        if not folder:
            self._set_status("Export canceled.")
            return
        for i, rows in enumerate(self.day_tables_data, start=1):
            path = os.path.join(folder, f"ymaze_day_{i}.csv")
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                for row in rows:
                    writer.writerow(row)
        self._write_output(f"Exported {len(self.day_tables_data)} CSVs to: {folder}\n")
        self._set_status("Export complete (CSVs).")

    def on_export_one_csv(self):
        if not self.day_tables_data:
            self._write_output("Nothing to export. Generate tables first.\n")
            self._set_status("Export aborted: nothing to export.")
            return
        save_path = filedialog.asksaveasfilename(
            title="Save combined CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", ".csv")],
            initialfile="ymaze_all_days.csv",
        )
        if not save_path:
            self._set_status("Export canceled.")
            return

        with open(save_path, "w", newline="") as f:
            writer = csv.writer(f)
            for day_idx, rows in enumerate(self.day_tables_data, start=1):
                header = ["Day"] + rows[0]
                if day_idx == 1:
                    writer.writerow(header)
                for row in rows[1:]:
                    writer.writerow([day_idx] + row)
        self._write_output(f"Exported combined CSV to: {save_path}\n")
        self._set_status("Export complete (combined CSV).")

    def on_export_xlsx(self):
        if not self.day_tables_data:
            self._write_output("Nothing to export. Generate tables first.\n")
            self._set_status("Export aborted: nothing to export.")
            return
        try:
            from openpyxl import Workbook
        except Exception:
            messagebox.showwarning(
                "openpyxl not available",
                "To export Excel with one sheet per day, install openpyxl:\n\n    pip install openpyxl\n\nThen try again."
            )
            self._set_status("Export aborted: openpyxl not installed.")
            return

        save_path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", ".xlsx")],
            initialfile="ymaze_schedule.xlsx",
        )
        if not save_path:
            self._set_status("Export canceled.")
            return

        wb = Workbook()
        wb.remove(wb.active)  # drop default sheet

        for day_idx, rows in enumerate(self.day_tables_data, start=1):
            ws = wb.create_sheet(title=f"Day {day_idx}")
            for r_i, row in enumerate(rows, start=1):
                for c_i, val in enumerate(row, start=1):
                    ws.cell(row=r_i, column=c_i, value=val)
        wb.save(save_path)
        self._write_output(f"Exported Excel with {len(self.day_tables_data)} sheet(s) to: {save_path}\n")
        self._set_status("Export complete (Excel).")

    # ---------- Helpers ---------- #
    def _write_output(self, msg: str):
        self.txt_output.insert(tk.END, msg)
        self.txt_output.see(tk.END)

    def _set_status(self, msg: str):
        self.var_status.set(msg)


def main():
    app = YMazeApp()
    app.mainloop()


if __name__ == "__main__":
    main()
